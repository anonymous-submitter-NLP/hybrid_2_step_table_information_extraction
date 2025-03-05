import openpyxl
from openpyxl.styles import PatternFill
import openpyxl
from openpyxl.styles import Font
from pathlib import Path
from openpyxl.utils.dataframe import dataframe_to_rows

import numpy as np
import pandas as pd
import inspect
import os
from typing import List, Dict, Union

def create_colored_excel(string_matrix: np.array, 
                         float_matrix: Union[List, np.array], 
                         params: dict = None, 
                         similarity_matrices: List[Dict] = None, 
                         filename="Tests/test_results/test_rule_based_extraction/temp.xlsx", 
                         folder_path_to_store_tables: str = None, 
                         df_scores: pd.DataFrame = None):
    """
    Create an Excel file with colored cells based on the provided matrices and parameters.
    """
    wb = openpyxl.Workbook()
    create_main_sheet(wb, string_matrix, float_matrix, params, similarity_matrices)
    if df_scores is not None:
        add_dataframe_sheet(wb, df_scores, "Scores")
    wb.save(filename)
    

def add_dataframe_sheet(wb, dataframe, sheet_name):
    """
    Add a new sheet to the workbook with the provided dataframe.
    """
    ws = wb.create_sheet(title=sheet_name)
    for r in dataframe_to_rows(dataframe, index=True, header=True):
        ws.append(r)

def create_main_sheet(wb, string_matrix, float_matrix, params, similarity_matrices):
    """
    Create the main sheet in the workbook with the provided matrices and parameters.
    """
    ws = wb.active
    ws.title = "Main Sheet"

    col1 = list(params.keys())
    col2 = list(params.values())

    rows, cols = len(string_matrix), len(string_matrix[0])
    min_val, max_val = np.min(float_matrix), np.max(float_matrix)
    norm_matrix = 255 * (float_matrix - min_val) / (max_val - min_val) if max_val - min_val != 0 else np.zeros_like(float_matrix)
    i_max_values = np.argwhere(float_matrix == np.amax(float_matrix))

    add_string_matrix(ws, string_matrix, norm_matrix, i_max_values)
    float_start_row = rows + 5
    add_float_matrix(ws, float_matrix, norm_matrix, i_max_values, float_start_row)
    add_params(ws, params, cols + 1)
    if similarity_matrices is not None:
        add_similarity_matrices(ws, similarity_matrices, float_start_row + rows + 2)

def add_string_matrix(ws, string_matrix, norm_matrix, i_max_values):
    """
    Add the string matrix to the worksheet with colored cells based on the normalized matrix.
    """
    rows, cols = len(string_matrix), len(string_matrix[0])
    for i in range(rows):
        for j in range(cols):
            cell = ws.cell(row=i+1, column=j+1, value=string_matrix[i][j])
            hex_color = "add8e6" if [i, j] in i_max_values.tolist() else f"{255-int(norm_matrix[i][j]):02X}FF{255-int(norm_matrix[i][j]):02X}"
            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def add_float_matrix(ws, float_matrix, norm_matrix, i_max_values, start_row):
    """
    Add the float matrix to the worksheet with colored cells based on the normalized matrix.
    """
    rows, cols = len(float_matrix), len(float_matrix[0])
    ws.cell(row=start_row - 1, column=1, value="Accepting Matrix")
    for i in range(rows):
        for j in range(cols):
            ws.cell(row=start_row + i, column=j+1, value=float_matrix[i][j])
            hex_color = "add8e6" if [i, j] in i_max_values.tolist() else f"{255-int(norm_matrix[i][j]):02X}FF{255-int(norm_matrix[i][j]):02X}"
            ws.cell(row=start_row + i, column=j+1).fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def add_params(ws, params, start_col):
    """
    Add the parameters to the worksheet.
    """
    col1 = list(params.keys())
    col2 = list(params.values())
    if col1:
        for i, value in enumerate(col1):
            ws.cell(row=i+1, column=start_col+3, value=value)
    if col2:
        for i, value in enumerate(col2):
            ws.cell(row=i+1, column=start_col+4, value=str(value))

def add_similarity_matrices(ws, similarity_matrices, start_row):
    """
    Add the similarity matrices to the worksheet with colored cells based on the maximum values.
    """
    for i, similarity_matrix in enumerate(similarity_matrices):
        ws.cell(row=start_row, column=1, value=f"Similarity Matrix {i}")
        ws.cell(row=start_row+1, column=1, value=f'Query: {similarity_matrix["query"]}')
        file_name = inspect.getfile(similarity_matrix["similarity_metric"])
        similarity_metric_name = os.path.basename(file_name).split(".")[0]
        ws.cell(row=start_row+2, column=1, value=f'similarity_metric: {similarity_metric_name}')

        for j in range(similarity_matrix["similarity_matrix"].shape[0]):
            for k in range(similarity_matrix["similarity_matrix"].shape[1]):
                temp_similarity_matrix = np.maximum(similarity_matrix["similarity_matrix"], 0)
                max_value = np.max(temp_similarity_matrix + [0.0001])
                hex_color = "add8e6" if temp_similarity_matrix[j][k] == max_value else f"{255-int(255 * (temp_similarity_matrix[j][k] / max_value)):02X}FF{255-int(255 * (temp_similarity_matrix[j][k] / max_value)):02X}"
                ws.cell(row=start_row+j+3, column=k+1).fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                ws.cell(row=start_row+j+3, column=k+1, value=similarity_matrix["similarity_matrix"][j][k])

        start_row += similarity_matrix["similarity_matrix"].shape[0] + 4

# Example usage
if __name__ == "__main__":
    strings = [["A", "B", "C"], ["D", "E", "F"], ["G", "H", "I"]]
    floats = np.array([[0.1, 0.5, 0.9], [0.2, 0.6, 1.0], [0.3, 0.7, 0.8]])
    params = {1:3, 2:1}
    create_colored_excel(strings, floats, params)